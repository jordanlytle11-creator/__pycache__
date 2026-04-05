from datetime import datetime, timedelta, date
from fastapi import FastAPI, Depends, HTTPException, status, Header, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import text, inspect as sa_inspect
from passlib.context import CryptContext
from jose import JWTError, jwt
from uuid import uuid4
from typing import List, Optional
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from pathlib import Path
from email.message import EmailMessage
import io
import json
import os
import re
import smtplib

from app import models
from app.db import SessionLocal, engine
from app.models import CrmRecord, User, Role, InviteToken, LinkControl
from app.schemas import (
    CrmRecordCreate,
    CrmRecordRead,
    UserCreate,
    UserRead,
    UserUpdate,
    InviteCreate,
    InviteAccept,
    LinkControlCreate,
    ForgotCredentialsRequest,
    ExcelSecondTabImportResult,
    WorkbookTabSummary,
    WorkbookImportResult,
    WorkbookTabRowsResult,
)

models.Base.metadata.create_all(bind=engine)

# ── Schema migration: add new lease columns to crm_records if missing ─────────
def _migrate_crm_columns():
    inspector = sa_inspect(engine)
    existing_cols = {c['name'] for c in inspector.get_columns('crm_records')}
    new_cols = [
        ('lease_agent',       'VARCHAR(255)'),
        ('lease_agent_notes', 'TEXT'),
        ('lessor_owner',      'VARCHAR(255)'),
        ('lessee',            'VARCHAR(255)'),
        ('lease_date',        'DATE'),
        ('vol',               'VARCHAR(50)'),
        ('pg',                'VARCHAR(50)'),
        ('tract_description', 'TEXT'),
        ('gross_acres',       'FLOAT'),
        ('net_acres',         'FLOAT'),
        ('royalty',           'VARCHAR(50)'),
        ('bonus_agreed',      'VARCHAR(50)'),
        ('term_months',       'INTEGER'),
        ('extension_months',  'INTEGER'),
        ('mailed_date',       'DATE'),
    ]
    with engine.connect() as conn:
        for col_name, col_type in new_cols:
            if col_name not in existing_cols:
                conn.execute(text(f'ALTER TABLE crm_records ADD COLUMN {col_name} {col_type}'))
        conn.commit()

_migrate_crm_columns()

app = FastAPI(title='Local ERP/CRM MVP')

FRONTEND_DIR = Path(__file__).resolve().parents[1] / "frontend"
FORGOT_CREDENTIALS_LOG = Path(__file__).resolve().parents[1] / "forgot_credentials_requests.log"
INVITE_REQUESTS_LOG = Path(__file__).resolve().parents[1] / "invite_requests.log"

# security utils
SECRET_KEY = os.getenv('LOCAL_ERP_SECRET_KEY', 'change-this-secret-in-production')
ALGORITHM = 'HS256'
ACCESS_TOKEN_EXPIRE_MINUTES = 60
pwd_context = CryptContext(schemes=['pbkdf2_sha256', 'bcrypt'], deprecated='auto')
oauth2_scheme = OAuth2PasswordBearer(tokenUrl='token')


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta if expires_delta else timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({'exp': expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_user_by_email(db: Session, email: str):
    return db.query(User).filter(User.email == email).first()


def authenticate_user(db: Session, email: str, password: str):
    user = get_user_by_email(db, email)
    if not user or not verify_password(password, user.hashed_password):
        return None
    return user


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Could not validate credentials', headers={'WWW-Authenticate': 'Bearer'})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get('sub')
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = get_user_by_email(db, email)
    if user is None:
        raise credentials_exception
    return user


def get_current_active_user(current_user: User = Depends(get_current_user)):
    if not current_user.is_active:
        raise HTTPException(status_code=400, detail='Inactive user')
    return current_user


def require_admin(user: User = Depends(get_current_active_user)):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin privileges required')
    return user


def get_link_control(token: str, db: Session):
    link = db.query(LinkControl).filter(LinkControl.token == token, LinkControl.is_active == True).first()
    if not link or (link.expires_at and link.expires_at < datetime.utcnow()):
        return None
    return link


def require_manager_or_admin(user: User = Depends(get_current_active_user)):
    if not (user.is_manager or user.is_admin):
        raise HTTPException(status_code=403, detail='Manager or Admin privileges required')
    return user


def require_link_permission(required_permission: str):
    def checker(x_access_link: str = Header(None), db: Session = Depends(get_db)):
        if not x_access_link:
            raise HTTPException(status_code=401, detail='Link token required')
        link = get_link_control(x_access_link, db)
        if not link:
            raise HTTPException(status_code=401, detail='Invalid or expired link token')
        if link.permission != required_permission:
            raise HTTPException(status_code=403, detail='Link does not grant required permission')
        return link
    return checker


def send_forgot_credentials_email_to_admin(subject: str, body: str) -> bool:
    smtp_host = os.getenv('LOCAL_ERP_SMTP_HOST')
    smtp_port = int(os.getenv('LOCAL_ERP_SMTP_PORT', '587'))
    smtp_user = os.getenv('LOCAL_ERP_SMTP_USER')
    smtp_password = os.getenv('LOCAL_ERP_SMTP_PASSWORD')
    smtp_from = os.getenv('LOCAL_ERP_SMTP_FROM') or smtp_user
    admin_email = os.getenv('LOCAL_ERP_ADMIN_EMAIL', 'admin@localerp.com')

    if not smtp_host or not smtp_from:
        return False

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_from
    msg['To'] = admin_email
    msg.set_content(body)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        server.starttls()
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.send_message(msg)

    return True


def send_invite_email_to_user(recipient_email: str, role: str, token: str, expires_at: datetime) -> bool:
    smtp_host = os.getenv('LOCAL_ERP_SMTP_HOST')
    smtp_port = int(os.getenv('LOCAL_ERP_SMTP_PORT', '587'))
    smtp_user = os.getenv('LOCAL_ERP_SMTP_USER')
    smtp_password = os.getenv('LOCAL_ERP_SMTP_PASSWORD')
    smtp_from = os.getenv('LOCAL_ERP_SMTP_FROM') or smtp_user
    app_url = os.getenv('LOCAL_ERP_APP_URL', 'https://local-erp.onrender.com').rstrip('/')

    if not smtp_host or not smtp_from:
        return False

    invite_link = f"{app_url}/?invite_token={token}"
    login_link = app_url
    expires_text = expires_at.strftime('%Y-%m-%d %H:%M:%S UTC')
    body = '\n'.join([
        'You were invited to Local ERP/CRM.',
        '',
        f'Role: {role}',
        f'Login: {login_link}',
        f'Invite link: {invite_link}',
        f'Invite token: {token}',
        f'Expires at: {expires_text}',
        '',
        'If the link does not auto-fill the token, open the app and paste the token in Accept Invite.',
    ])

    msg = EmailMessage()
    msg['Subject'] = 'You are invited to Local ERP/CRM'
    msg['From'] = smtp_from
    msg['To'] = recipient_email
    msg.set_content(body)
    msg.add_alternative(
        f"""
<html>
  <body>
    <p>You were invited to <b>Local ERP/CRM</b>.</p>
    <p><b>Role:</b> {role}</p>
    <p><a href=\"{login_link}\">Open Login</a></p>
    <p><a href=\"{invite_link}\">Accept Invite</a></p>
    <p><b>Invite token:</b> {token}</p>
    <p><b>Expires at:</b> {expires_text}</p>
    <p>If the invite link does not auto-fill the token, copy/paste the token in the Accept Invite section.</p>
  </body>
</html>
""",
        subtype='html',
    )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        server.starttls()
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.send_message(msg)

    return True


def _normalize_column_name(raw_header: str, existing: set[str]) -> str:
    cleaned = re.sub(r'[^a-zA-Z0-9]+', '_', str(raw_header).strip().lower()).strip('_')
    if not cleaned:
        cleaned = 'column'
    if cleaned[0].isdigit():
        cleaned = f'col_{cleaned}'
    candidate = cleaned
    i = 2
    while candidate in existing:
        candidate = f'{cleaned}_{i}'
        i += 1
    existing.add(candidate)
    return candidate


def _normalize_table_name(raw_name: str, index: int) -> str:
    base = re.sub(r'[^a-zA-Z0-9]+', '_', str(raw_name).strip().lower()).strip('_')
    if not base:
        base = f'sheet_{index}'
    if base[0].isdigit():
        base = f'sheet_{base}'
    return f'tomahawk_{base}'


def _stringify_cell(value) -> Optional[str]:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _row_has_values(values: list[object]) -> bool:
    return any(_stringify_cell(value) for value in values)


def _header_score(values: list[object]) -> int:
    normalized_values = [
        re.sub(r'[^a-z0-9]+', '_', value.lower()).strip('_')
        for value in (_stringify_cell(item) for item in values)
        if value
    ]

    trs_tokens = {'township', 'range', 'section', 't_r_s', 'trs', 'town_range', 'twp', 'twn', 'rng', 'sec'}
    trs_hits = sum(1 for v in normalized_values if v in trs_tokens or any(k in v for k in ('township', 'section')))

    # Rows with fewer than 2 non-empty cells are not headers unless they contain TRS keywords
    if len(normalized_values) < 2 and trs_hits == 0:
        return -1

    known_tokens = {
        'ami_aoi',
        'state_code',
        'county_code',
        'state',
        'county',
        't_r_s',
        'town_range',
        'township',
        'range',
        'section',
        'location',
        'location_number',
        'well_name',
        'dsu_name',
        'pad_name',
        'lease_name',
        'lease_number',
        'interest_type',
        'property_id',
        'assessor_pin',
    }
    score = len(normalized_values) * 2
    for token in normalized_values:
        if token in known_tokens:
            score += 8
        elif any(keyword in token for keyword in ('lease', 'owner', 'town', 'range', 'section', 'well', 'assessor', 'property')):
            score += 3

    # Strong bonus for rows that contain TRS field names — these are almost always real header rows
    if trs_hits >= 2:
        score += 100
    elif trs_hits == 1:
        score += 30

    return score


def _detect_header_row(sheet) -> tuple[int, list[object]]:
    max_scan_rows = min(sheet.max_row, 25)
    best_score = -1
    best_row_index = 1
    best_values: list[object] = []

    for row_index in range(1, max_scan_rows + 1):
        values = list(next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True)))
        score = _header_score(values)
        if score < 0:
            continue

        lookahead_score = 0
        for next_row_index in range(row_index + 1, min(sheet.max_row, row_index + 3) + 1):
            next_values = list(next(sheet.iter_rows(min_row=next_row_index, max_row=next_row_index, values_only=True)))
            non_empty = sum(1 for value in next_values if _stringify_cell(value))
            lookahead_score += min(non_empty, 6)

        total_score = score + lookahead_score
        if total_score > best_score:
            best_score = total_score
            best_row_index = row_index
            best_values = values

    if not best_values:
        raise HTTPException(status_code=400, detail=f'Could not detect header row in sheet: {sheet.title}')

    return best_row_index, best_values


def _first_value(row_map: dict[str, Optional[str]], aliases: list[str]) -> Optional[str]:
    for alias in aliases:
        value = row_map.get(alias)
        if value:
            return value
    return None


def _lease_sheet_aliases(sheet_name: str, primary: list[str], fallback: list[str]) -> list[str]:
    normalized_sheet = re.sub(r'[^a-z0-9]+', ' ', sheet_name.lower()).strip()
    if 'lease acquisitions' in normalized_sheet or 'lacq' in normalized_sheet:
        return primary + fallback
    return primary


def _parse_int_token(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    match = re.search(r'(\d+)', value)
    return int(match.group(1)) if match else None


def _parse_float_token(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    cleaned = re.sub(r'[^0-9.\-]+', '', value)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_date_token(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    text_value = value.strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_trs_components(row_map: dict[str, Optional[str]]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    section = _parse_int_token(_first_value(row_map, ['section', 'sec', 'sect', 'column_20']))
    township = _parse_int_token(_first_value(row_map, ['township', 'twp', 'twn', 'town', 'column_18']))
    range_value = _parse_int_token(_first_value(row_map, ['range', 'rng', 'rge', 'column_19']))

    trs_text = _first_value(row_map, ['t_r_s', 'trs', 'town_range'])
    if trs_text:
        trs_match = re.search(r'(\d+)\s*[NSEW]?[-\s]+(\d+)\s*[NSEW]?[-\s]+(\d+)', trs_text, flags=re.IGNORECASE)
        if trs_match:
            township = township or int(trs_match.group(1))
            range_value = range_value or int(trs_match.group(2))
            section = section or int(trs_match.group(3))
        else:
            town_range_match = re.search(r'(\d+)\s*[NSEW]?[-\s]+(\d+)\s*[NSEW]?', trs_text, flags=re.IGNORECASE)
            if town_range_match:
                township = township or int(town_range_match.group(1))
                range_value = range_value or int(town_range_match.group(2))

    town_range_text = _first_value(row_map, ['town_range'])
    if town_range_text:
        town_range_match = re.search(r'(\d+)\s*[NSEW]?[-\s]+(\d+)\s*[NSEW]?', town_range_text, flags=re.IGNORECASE)
        if town_range_match:
            township = township or int(town_range_match.group(1))
            range_value = range_value or int(town_range_match.group(2))

    return township, range_value, section


def _resolve_workbook_status(row_map: dict[str, Optional[str]]) -> Optional[str]:
    status_value = _first_value(row_map, ['status', 'lease_status', 'crm_status', 'current_status', 'column_22'])
    if status_value:
        return status_value

    for key, value in row_map.items():
        if value and 'status' in key:
            return value

    return None


def _build_crm_record_payload(sheet_name: str, tab_key: str, source_row_number: int, row_map: dict[str, Optional[str]]) -> Optional[dict]:
    township, range_value, section = _parse_trs_components(row_map)
    if township is None or range_value is None or section is None:
        return None

    company = _first_value(row_map, _lease_sheet_aliases(sheet_name, [
        'lease_name',
        'owner_name',
        'owner',
        'company',
        'well_name',
        'dsu_name',
        'pad_name',
        'property_id',
        'assessor_pin',
        'location_number',
    ], ['column_13', 'column_10', 'column_14']))
    contact = _first_value(row_map, _lease_sheet_aliases(sheet_name, [
        'contact',
        'owner_name',
        'lease_name',
        'owner',
        'well_name',
        'property_id',
    ], ['column_13', 'column_10']))
    status_value = _resolve_workbook_status(row_map) or 'No Contact'

    lease_agent = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['lease_agent', 'landman', 'agent'], ['column_29']))
    lease_agent_notes = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['lease_agent_notes', 'agent_notes', 'notes', 'remarks', 'comment'], ['column_31']))
    lessor_owner = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['lessor_owner', 'owner_name', 'owner', 'lessor'], ['column_13', 'column_10']))
    lessee = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['lessee', 'operator', 'company'], ['column_14']))
    lease_date = _parse_date_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['lease_date', 'effective_date'], ['column_15'])))
    vol = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['vol', 'volume'], ['column_16']))
    pg = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['pg', 'page'], ['column_17']))
    tract_description = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['tract_description', 'legal_description', 'description', 'tract'], ['column_21']))
    gross_acres = _parse_float_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['gross_acres', 'gross_acreage', 'acres_gross'], ['column_23'])))
    net_acres = _parse_float_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['net_acres', 'net_acreage', 'acres_net'], ['column_24'])))
    royalty = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['royalty', 'royalty_rate'], ['column_25']))
    bonus_agreed = _first_value(row_map, _lease_sheet_aliases(sheet_name, ['bonus_agreed', 'bonus'], ['column_26']))
    term_months = _parse_int_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['term_month', 'term_months', 'term'], ['column_27'])))
    extension_months = _parse_int_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['extension_month', 'extension_months', 'extension'], ['column_28'])))
    mailed_date = _parse_date_token(_first_value(row_map, _lease_sheet_aliases(sheet_name, ['mailed_date', 'mail_date', 'date_mailed'], ['column_30'])))

    trs_text = _first_value(row_map, ['t_r_s']) or f'{township}-{range_value}-{section}'
    fallback_label = f'{sheet_name} {trs_text}'
    company = company or fallback_label
    contact = contact or company

    extra_data = {
        key: value
        for key, value in row_map.items()
        if value and key not in {'lease_name', 'owner_name', 'owner', 'company', 'contact', 'well_name'}
    }
    extra_data.update({
        'workbook_sheet': sheet_name,
        'workbook_tab_key': tab_key,
        'workbook_source_row': source_row_number,
    })

    return {
        'company': company,
        'contact': contact,
        'status': status_value,
        'township': township,
        'range': range_value,
        'section': section,
        'lease_agent': lease_agent,
        'lease_agent_notes': lease_agent_notes,
        'lessor_owner': lessor_owner,
        'lessee': lessee,
        'lease_date': lease_date,
        'vol': vol,
        'pg': pg,
        'tract_description': tract_description,
        'gross_acres': gross_acres,
        'net_acres': net_acres,
        'royalty': royalty,
        'bonus_agreed': bonus_agreed,
        'term_months': term_months,
        'extension_months': extension_months,
        'mailed_date': mailed_date,
        'extra_data': extra_data,
    }


def _normalize_company_key(company: Optional[str]) -> str:
    if not company:
        return ''
    return re.sub(r'\s+', ' ', company).strip().lower()


def _crm_record_key(payload: dict) -> str:
    company_key = _normalize_company_key(payload.get('company'))
    township = payload.get('township')
    range_value = payload.get('range')
    section = payload.get('section')
    return f'{company_key}|{township}|{range_value}|{section}'


def _merge_crm_payload(current: dict, incoming: dict) -> dict:
    merged = current.copy()

    for field in (
        'company', 'contact', 'status', 'township', 'range', 'section',
        'lease_agent', 'lease_agent_notes', 'lessor_owner', 'lessee', 'lease_date',
        'vol', 'pg', 'tract_description', 'gross_acres', 'net_acres', 'royalty',
        'bonus_agreed', 'term_months', 'extension_months', 'mailed_date',
    ):
        value = incoming.get(field)
        if value is not None and value != '':
            merged[field] = value

    current_extra = merged.get('extra_data') or {}
    incoming_extra = incoming.get('extra_data') or {}
    merged['extra_data'] = {**current_extra, **incoming_extra}

    return merged


@app.post('/token')
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=400, detail='Incorrect username or password')
    role = 'admin' if user.is_admin else ('manager' if user.is_manager else 'employee')
    access_token = create_access_token(data={'sub': user.email, 'role': role})
    return {'access_token': access_token, 'token_type': 'bearer'}


@app.post('/admin/users', response_model=UserRead)
def create_user(user_in: UserCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    if get_user_by_email(db, user_in.email):
        raise HTTPException(status_code=400, detail='Email already registered')
    hashed = get_password_hash(user_in.password)
    user = User(email=user_in.email, hashed_password=hashed, is_admin=(user_in.role == 'admin'), is_manager=(user_in.role == 'manager'))
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.get('/admin/users', response_model=List[UserRead])
def list_users(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    users = db.query(User).all()
    return [
        UserRead(
            id=user.id,
            email=user.email,
            is_active=user.is_active,
            is_manager=user.is_manager,
            is_admin=user.is_admin,
            role='admin' if user.is_admin else ('manager' if user.is_manager else 'employee'),
        )
        for user in users
    ]


@app.patch('/admin/users/{user_id}', response_model=UserRead)
def update_user(user_id: int, update: UserUpdate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')
    if update.role is not None:
        user.is_admin = update.role == 'admin'
        user.is_manager = update.role == 'manager'
    if update.password is not None:
        if len(update.password) < 8:
            raise HTTPException(status_code=400, detail='Password must be at least 8 characters')
        user.hashed_password = get_password_hash(update.password)
    db.commit()
    db.refresh(user)
    return UserRead(
        id=user.id,
        email=user.email,
        is_active=user.is_active,
        is_manager=user.is_manager,
        is_admin=user.is_admin,
        role='admin' if user.is_admin else ('manager' if user.is_manager else 'employee'),
    )


@app.post('/admin/invite')
def create_invite(invite: InviteCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    token = str(uuid4())
    expiry = datetime.utcnow() + timedelta(days=7)
    invite_obj = InviteToken(token=token, email=invite.email, role=invite.role, expires_at=expiry)
    db.add(invite_obj)
    db.commit()

    try:
        sent = send_invite_email_to_user(invite.email, invite.role, token, expiry)
    except Exception:
        sent = False

    app_url = os.getenv('LOCAL_ERP_APP_URL', 'https://local-erp.onrender.com').rstrip('/')
    invite_link = f"{app_url}/?invite_token={token}"

    if not sent:
        now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        with open(INVITE_REQUESTS_LOG, 'a', encoding='utf-8') as f:
            f.write('\n'.join([
                'Invite email delivery fallback (not sent via SMTP).',
                f'Timestamp: {now}',
                f'Recipient: {invite.email}',
                f'Role: {invite.role}',
                f'Invite link: {invite_link}',
                f'Invite token: {token}',
                f'Expires at: {expiry.isoformat()}',
                '-' * 72,
            ]) + '\n')

    return {
        'invite_token': token,
        'invite_link': invite_link,
        'expires_at': expiry.isoformat(),
        'role': invite.role,
        'delivery': 'email' if sent else 'logged',
    }


@app.post('/admin/accept-invite')
def accept_invite(invitation: InviteAccept, db: Session = Depends(get_db)):
    invite = db.query(InviteToken).filter(InviteToken.token == invitation.token, InviteToken.is_used == False).first()
    if not invite or invite.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail='Invalid or expired invite token')

    if get_user_by_email(db, invite.email):
        raise HTTPException(status_code=400, detail='User already exists')

    hashed = get_password_hash(invitation.password)
    user = User(email=invite.email, hashed_password=hashed, is_active=True,
                is_admin=(invite.role == 'admin'), is_manager=(invite.role == 'manager'))
    db.add(user)
    invite.is_used = True
    db.commit()
    db.refresh(user)
    return {'message': 'User created', 'email': user.email, 'role': invite.role}


@app.post('/auth/forgot-credentials')
def forgot_credentials(request: ForgotCredentialsRequest):
    if not request.email and not request.username:
        raise HTTPException(status_code=400, detail='Provide an email or username')

    admin_email = os.getenv('LOCAL_ERP_ADMIN_EMAIL', 'admin@localerp.com')
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
    lines = [
        'A user submitted a forgot username/password request.',
        '',
        f'Timestamp: {now}',
        f'Email provided: {request.email or "(not provided)"}',
        f'Username provided: {request.username or "(not provided)"}',
        f'Note: {request.message or "(none)"}',
    ]
    body = '\n'.join(lines)

    try:
        sent = send_forgot_credentials_email_to_admin(
            subject='Local ERP forgot credentials request',
            body=body,
        )
    except Exception:
        sent = False

    if not sent:
        with open(FORGOT_CREDENTIALS_LOG, 'a', encoding='utf-8') as f:
            f.write(body + '\n' + ('-' * 72) + '\n')

    return {
        'message': 'Request submitted',
        'notified_admin_email': admin_email,
        'delivery': 'email' if sent else 'logged',
    }


@app.post('/admin/import-excel-second-tab', response_model=ExcelSecondTabImportResult)
def import_excel_second_tab_to_db(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl is required for Excel import')

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail='Uploaded file is empty')

    try:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f'Invalid Excel file: {exc}')

    if len(workbook.worksheets) < 2:
        raise HTTPException(status_code=400, detail='Workbook must have at least two tabs')

    sheet = workbook.worksheets[1]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    if not any(h is not None and str(h).strip() for h in headers):
        raise HTTPException(status_code=400, detail='Second tab has no header row')

    normalized_names: list[str] = []
    seen_names: set[str] = set()
    for idx, header in enumerate(headers, start=1):
        header_value = str(header).strip() if header is not None else f'column_{idx}'
        normalized_names.append(_normalize_column_name(header_value, seen_names))

    table_name = 'tomahawk_second_tab_rows'
    created_columns: list[str] = []
    imported_rows = 0

    with engine.begin() as conn:
        conn.execute(text(f'CREATE TABLE IF NOT EXISTS {table_name} (id INTEGER PRIMARY KEY AUTOINCREMENT)'))

        existing_cols = {
            row[1]
            for row in conn.execute(text(f'PRAGMA table_info({table_name})')).fetchall()
        }

        for col in normalized_names:
            if col not in existing_cols:
                conn.execute(text(f'ALTER TABLE {table_name} ADD COLUMN "{col}" TEXT'))
                created_columns.append(col)

        insert_cols = ', '.join([f'"{col}"' for col in normalized_names])
        insert_placeholders = ', '.join([f':{col}' for col in normalized_names])
        insert_sql = text(f'INSERT INTO {table_name} ({insert_cols}) VALUES ({insert_placeholders})')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            payload = {}
            for i, col in enumerate(normalized_names):
                val = row[i] if i < len(row) else None
                payload[col] = None if val is None else str(val)
            conn.execute(insert_sql, payload)
            imported_rows += 1

    return ExcelSecondTabImportResult(
        table_name=table_name,
        sheet_name=sheet.title,
        headers_detected=len(normalized_names),
        columns_created=len(created_columns),
        rows_imported=imported_rows,
        created_columns=created_columns,
    )


@app.post('/admin/import-excel-workbook', response_model=WorkbookImportResult)
def import_excel_workbook(
    file: UploadFile = File(...),
    user: User = Depends(require_manager_or_admin),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl is required for Excel import')

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail='Uploaded file is empty')

    try:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f'Invalid Excel file: {exc}')

    workbook_name = file.filename or 'uploaded_workbook.xlsx'

    tabs: list[WorkbookTabSummary] = []
    total_rows = 0
    crm_rows_by_key: dict[str, dict] = {}

    with engine.begin() as conn:
        conn.execute(text('''
            CREATE TABLE IF NOT EXISTS workbook_tabs_index (
                tab_key TEXT PRIMARY KEY,
                workbook_name TEXT NOT NULL DEFAULT '',
                sheet_name TEXT NOT NULL,
                table_name TEXT NOT NULL,
                headers_json TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                imported_at TEXT NOT NULL
            )
        '''))
        idx_columns = {
            row[1]
            for row in conn.execute(text('PRAGMA table_info(workbook_tabs_index)')).fetchall()
        }
        if 'workbook_name' not in idx_columns:
            conn.execute(text("ALTER TABLE workbook_tabs_index ADD COLUMN workbook_name TEXT NOT NULL DEFAULT ''"))
        conn.execute(text('''
            CREATE TABLE IF NOT EXISTS workbook_crm_import_index (
                tab_key TEXT NOT NULL,
                source_row_number INTEGER NOT NULL,
                crm_record_id INTEGER NOT NULL,
                imported_at TEXT NOT NULL,
                PRIMARY KEY (tab_key, source_row_number)
            )
        '''))

        for idx, sheet in enumerate(workbook.worksheets, start=1):
            if sheet.max_row < 1:
                continue

            header_row_number, headers_raw = _detect_header_row(sheet)

            normalized_names: list[str] = []
            seen_names: set[str] = set()
            for col_idx, header in enumerate(headers_raw, start=1):
                header_value = str(header).strip() if header is not None else f'column_{col_idx}'
                normalized_names.append(_normalize_column_name(header_value, seen_names))

            tab_key = f'tab_{idx}'
            table_name = _normalize_table_name(sheet.title, idx)

            conn.execute(text(f'CREATE TABLE IF NOT EXISTS "{table_name}" (id INTEGER PRIMARY KEY AUTOINCREMENT, source_row_number INTEGER)'))

            existing_cols = {
                row[1]
                for row in conn.execute(text(f'PRAGMA table_info("{table_name}")')).fetchall()
            }

            for col in normalized_names:
                if col not in existing_cols:
                    conn.execute(text(f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" TEXT'))

            conn.execute(text(f'DELETE FROM "{table_name}"'))
            insert_cols = ', '.join(['source_row_number'] + [f'"{col}"' for col in normalized_names])
            insert_placeholders = ', '.join([':source_row_number'] + [f':{col}' for col in normalized_names])
            insert_sql = text(f'INSERT INTO "{table_name}" ({insert_cols}) VALUES ({insert_placeholders})')
            imported_rows = 0

            for source_row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
                start=header_row_number + 1,
            ):
                row_values = list(row)
                if not _row_has_values(row_values):
                    continue

                payload = {'source_row_number': source_row_number}
                row_map: dict[str, Optional[str]] = {}
                for col_index, col in enumerate(normalized_names):
                    cell_value = row_values[col_index] if col_index < len(row_values) else None
                    normalized_value = _stringify_cell(cell_value)
                    payload[col] = normalized_value
                    row_map[col] = normalized_value

                conn.execute(insert_sql, payload)
                imported_rows += 1

                crm_payload = _build_crm_record_payload(sheet.title, tab_key, source_row_number, row_map)
                if crm_payload:
                    crm_key = _crm_record_key(crm_payload)
                    if crm_key:
                        if crm_key not in crm_rows_by_key:
                            crm_rows_by_key[crm_key] = {
                                'record': crm_payload,
                                'sources': [(tab_key, source_row_number)],
                            }
                        else:
                            crm_rows_by_key[crm_key]['record'] = _merge_crm_payload(
                                crm_rows_by_key[crm_key]['record'],
                                crm_payload,
                            )
                            crm_rows_by_key[crm_key]['sources'].append((tab_key, source_row_number))

            imported_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
            conn.execute(text('''
                INSERT INTO workbook_tabs_index (tab_key, workbook_name, sheet_name, table_name, headers_json, row_count, imported_at)
                VALUES (:tab_key, :workbook_name, :sheet_name, :table_name, :headers_json, :row_count, :imported_at)
                ON CONFLICT(tab_key) DO UPDATE SET
                    workbook_name=excluded.workbook_name,
                    sheet_name=excluded.sheet_name,
                    table_name=excluded.table_name,
                    headers_json=excluded.headers_json,
                    row_count=excluded.row_count,
                    imported_at=excluded.imported_at
            '''), {
                'tab_key': tab_key,
                'workbook_name': workbook_name,
                'sheet_name': sheet.title,
                'table_name': table_name,
                'headers_json': json.dumps(normalized_names),
                'row_count': imported_rows,
                'imported_at': imported_at,
            })

            total_rows += imported_rows
            tabs.append(WorkbookTabSummary(
                tab_key=tab_key,
                workbook_name=workbook_name,
                sheet_name=sheet.title,
                table_name=table_name,
                row_count=imported_rows,
                headers=normalized_names,
            ))

    if not tabs:
        raise HTTPException(status_code=400, detail='No usable tabs found (missing headers)')

    try:
        db.execute(text('''
            CREATE TABLE IF NOT EXISTS workbook_crm_import_index (
                tab_key TEXT NOT NULL,
                source_row_number INTEGER NOT NULL,
                crm_record_id INTEGER NOT NULL,
                imported_at TEXT NOT NULL,
                PRIMARY KEY (tab_key, source_row_number)
            )
        '''))
        db.execute(text('DELETE FROM workbook_crm_import_index'))

        imported_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        existing_records = db.query(CrmRecord).all()
        existing_by_key: dict[str, CrmRecord] = {}
        for existing in existing_records:
            key = _crm_record_key({
                'company': existing.company,
                'township': existing.township,
                'range': existing.range,
                'section': existing.section,
            })
            if not key:
                continue
            existing_by_key[key] = existing

        crm_records_imported = 0
        for crm_key, crm_row in crm_rows_by_key.items():
            record_payload = crm_row['record']
            existing = existing_by_key.get(crm_key)

            if existing:
                merged_extra = {**(existing.extra_data or {}), **(record_payload['extra_data'] or {})}
                existing.company = record_payload['company']
                existing.contact = record_payload['contact']
                existing.status = record_payload['status']
                existing.township = record_payload['township']
                existing.range = record_payload['range']
                existing.section = record_payload['section']
                existing.trscode = f"T{record_payload['township']}R{record_payload['range']}S{record_payload['section']}"
                existing.lease_agent = record_payload.get('lease_agent')
                existing.lease_agent_notes = record_payload.get('lease_agent_notes')
                existing.lessor_owner = record_payload.get('lessor_owner')
                existing.lessee = record_payload.get('lessee')
                existing.lease_date = record_payload.get('lease_date')
                existing.vol = record_payload.get('vol')
                existing.pg = record_payload.get('pg')
                existing.tract_description = record_payload.get('tract_description')
                existing.gross_acres = record_payload.get('gross_acres')
                existing.net_acres = record_payload.get('net_acres')
                existing.royalty = record_payload.get('royalty')
                existing.bonus_agreed = record_payload.get('bonus_agreed')
                existing.term_months = record_payload.get('term_months')
                existing.extension_months = record_payload.get('extension_months')
                existing.mailed_date = record_payload.get('mailed_date')
                existing.extra_data = merged_extra
                db_record = existing
            else:
                db_record = CrmRecord(
                    company=record_payload['company'],
                    contact=record_payload['contact'],
                    status=record_payload['status'],
                    township=record_payload['township'],
                    range=record_payload['range'],
                    section=record_payload['section'],
                    trscode=f"T{record_payload['township']}R{record_payload['range']}S{record_payload['section']}",
                    lease_agent=record_payload.get('lease_agent'),
                    lease_agent_notes=record_payload.get('lease_agent_notes'),
                    lessor_owner=record_payload.get('lessor_owner'),
                    lessee=record_payload.get('lessee'),
                    lease_date=record_payload.get('lease_date'),
                    vol=record_payload.get('vol'),
                    pg=record_payload.get('pg'),
                    tract_description=record_payload.get('tract_description'),
                    gross_acres=record_payload.get('gross_acres'),
                    net_acres=record_payload.get('net_acres'),
                    royalty=record_payload.get('royalty'),
                    bonus_agreed=record_payload.get('bonus_agreed'),
                    term_months=record_payload.get('term_months'),
                    extension_months=record_payload.get('extension_months'),
                    mailed_date=record_payload.get('mailed_date'),
                    extra_data=record_payload['extra_data'],
                )
                db.add(db_record)
                db.flush()
                existing_by_key[crm_key] = db_record

            for tab_key, source_row_number in crm_row['sources']:
                db.execute(text('''
                    INSERT INTO workbook_crm_import_index (tab_key, source_row_number, crm_record_id, imported_at)
                    VALUES (:tab_key, :source_row_number, :crm_record_id, :imported_at)
                '''), {
                    'tab_key': tab_key,
                    'source_row_number': source_row_number,
                    'crm_record_id': db_record.id,
                    'imported_at': imported_at,
                })

            crm_records_imported += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    return WorkbookImportResult(
        workbook_name=workbook_name,
        tabs_imported=len(tabs),
        total_rows_imported=total_rows,
        crm_records_imported=crm_records_imported,
        tabs=tabs,
    )


@app.get('/crm/workbook-tabs', response_model=list[WorkbookTabSummary])
def list_workbook_tabs(user: User = Depends(require_manager_or_admin)):
    rows: list[WorkbookTabSummary] = []
    with engine.begin() as conn:
        conn.execute(text('''
            CREATE TABLE IF NOT EXISTS workbook_tabs_index (
                tab_key TEXT PRIMARY KEY,
                workbook_name TEXT NOT NULL DEFAULT '',
                sheet_name TEXT NOT NULL,
                table_name TEXT NOT NULL,
                headers_json TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                imported_at TEXT NOT NULL
            )
        '''))
        idx_columns = {
            row[1]
            for row in conn.execute(text('PRAGMA table_info(workbook_tabs_index)')).fetchall()
        }
        if 'workbook_name' not in idx_columns:
            conn.execute(text("ALTER TABLE workbook_tabs_index ADD COLUMN workbook_name TEXT NOT NULL DEFAULT ''"))

        result = conn.execute(text('SELECT tab_key, workbook_name, sheet_name, table_name, headers_json, row_count FROM workbook_tabs_index ORDER BY tab_key')).fetchall()
        for row in result:
            rows.append(WorkbookTabSummary(
                tab_key=row[0],
                workbook_name=row[1] or row[2],
                sheet_name=row[2],
                table_name=row[3],
                row_count=row[5],
                headers=json.loads(row[4]),
            ))
    return rows


@app.get('/crm/workbook-tabs/{tab_key}/rows', response_model=WorkbookTabRowsResult)
def get_workbook_tab_rows(tab_key: str, limit: int = 200, user: User = Depends(require_manager_or_admin)):
    safe_limit = max(1, min(limit, 300000))

    with engine.begin() as conn:
        idx = conn.execute(text('SELECT sheet_name, table_name, headers_json FROM workbook_tabs_index WHERE tab_key = :tab_key'), {'tab_key': tab_key}).fetchone()
        if not idx:
            raise HTTPException(status_code=404, detail='Workbook tab not found')

        sheet_name = idx[0]
        table_name = idx[1]
        headers = json.loads(idx[2])
        select_cols = ', '.join([f'"{h}"' for h in headers])
        data_rows = conn.execute(text(f'SELECT id, source_row_number, {select_cols} FROM "{table_name}" LIMIT :limit'), {'limit': safe_limit}).fetchall()

        rows: list[dict] = []
        for data_row in data_rows:
            payload = {
                'id': data_row[0],
                'source_row_number': data_row[1],
            }
            for i, header in enumerate(headers, start=2):
                payload[header] = data_row[i]
            rows.append(payload)

    return WorkbookTabRowsResult(
        tab_key=tab_key,
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
    )


@app.post('/admin/link', response_model=LinkControlCreate)
def create_link_control(link_in: LinkControlCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    token = str(uuid4())
    expiry = datetime.utcnow() + timedelta(hours=link_in.expires_in_hours)
    link = LinkControl(token=token, permission=link_in.permission, expires_at=expiry)
    db.add(link)
    db.commit()
    return {'permission': link.permission, 'expires_in_hours': link_in.expires_in_hours}


@app.post('/crm', response_model=CrmRecordRead)
def create_record(record: CrmRecordCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    trscode = f"T{record.township}R{record.range}S{record.section}"
    db_record = CrmRecord(
        company=record.company, contact=record.contact, status=record.status,
        township=record.township, range=record.range, section=record.section,
        trscode=trscode, extra_data=record.extra_data,
        lease_agent=record.lease_agent, lease_agent_notes=record.lease_agent_notes,
        lessor_owner=record.lessor_owner, lessee=record.lessee,
        lease_date=record.lease_date, vol=record.vol, pg=record.pg,
        tract_description=record.tract_description,
        gross_acres=record.gross_acres, net_acres=record.net_acres,
        royalty=record.royalty, bonus_agreed=record.bonus_agreed,
        term_months=record.term_months, extension_months=record.extension_months,
        mailed_date=record.mailed_date,
    )
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    return db_record


@app.post('/crm/link-create', response_model=CrmRecordRead, dependencies=[Depends(require_link_permission('create_crm'))])
def create_record_link(record: CrmRecordCreate, db: Session = Depends(get_db)):
    trscode = f"T{record.township}R{record.range}S{record.section}"
    db_record = CrmRecord(
        company=record.company, contact=record.contact, status=record.status,
        township=record.township, range=record.range, section=record.section,
        trscode=trscode, extra_data=record.extra_data,
        lease_agent=record.lease_agent, lease_agent_notes=record.lease_agent_notes,
        lessor_owner=record.lessor_owner, lessee=record.lessee,
        lease_date=record.lease_date, vol=record.vol, pg=record.pg,
        tract_description=record.tract_description,
        gross_acres=record.gross_acres, net_acres=record.net_acres,
        royalty=record.royalty, bonus_agreed=record.bonus_agreed,
        term_months=record.term_months, extension_months=record.extension_months,
        mailed_date=record.mailed_date,
    )
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    return db_record


@app.get('/crm', response_model=List[CrmRecordRead])
def list_records(skip: int = 0, limit: int = 5000, db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    safe_limit = max(1, min(limit, 300000))
    return db.query(CrmRecord).offset(skip).limit(safe_limit).all()


@app.get('/crm/link-search', response_model=List[CrmRecordRead], dependencies=[Depends(require_link_permission('read_crm'))])
def search_records_link(township: int = None, range: int = None, section: int = None, status: str = None, db: Session = Depends(get_db)):
    q = db.query(CrmRecord)
    if township is not None:
        q = q.filter(CrmRecord.township == township)
    if range is not None:
        q = q.filter(CrmRecord.range == range)
    if section is not None:
        q = q.filter(CrmRecord.section == section)
    if status is not None:
        q = q.filter(CrmRecord.status == status)
    return q.limit(200).all()


@app.get('/crm/search', response_model=List[CrmRecordRead])
def search_records(
    township: int = None,
    range: int = None,
    section: int = None,
    status: str = None,
    limit: int = 5000,
    db: Session = Depends(get_db),
    user: User = Depends(require_manager_or_admin),
):
    q = db.query(CrmRecord)
    if township is not None:
        q = q.filter(CrmRecord.township == township)
    if range is not None:
        q = q.filter(CrmRecord.range == range)
    if section is not None:
        q = q.filter(CrmRecord.section == section)
    if status is not None:
        q = q.filter(CrmRecord.status == status)
    safe_limit = max(1, min(limit, 300000))
    return q.limit(safe_limit).all()


@app.post('/crm/shapefile')
def generate_shapefile(db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    try:
        import geopandas as gpd
        from shapely.geometry import Point
    except ImportError:
        raise HTTPException(status_code=500, detail='geopandas and shapely required')

    records = db.query(CrmRecord).all()
    gdf = gpd.GeoDataFrame(
        [{'company': r.company, 'trscode': r.trscode, 'status': r.status, 'geometry': Point(r.township, r.range)} for r in records],
        crs='EPSG:4326'
    )
    path = 'C:/Users/JordanLytle/Jordanlytle11 Dropbox/local-erp-crm-shapefile.shp'
    gdf.to_file(path)
    return {'shapefile_path': path, 'count': len(records)}


@app.post('/sharepoint/export')
def export_to_sharepoint(db: Session = Depends(get_db), user: User = Depends(require_admin)):
    # Placeholder: replace with actual SharePoint/Graph integration
    file_path = 'C:/Users/JordanLytle/OneDrive - Vaquero Resources LLC/local_erp_export.csv'
    records = db.query(CrmRecord).all()
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write('id,company,contact,status,township,range,section,trscode\n')
        for r in records:
            f.write(f"{r.id},{r.company},{r.contact},{r.status},{r.township},{r.range},{r.section},{r.trscode}\n")
    return {'sharepoint_path': file_path, 'detail': 'ready for SharePoint sync via OneDrive folder'}


@app.get('/manager/kpis')
def manager_kpis(db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    total = db.query(CrmRecord).count()
    statuses = {
        st: db.query(CrmRecord).filter(CrmRecord.status == st).count()
        for st in ['No Contact', 'Working', 'Signed / In Hand']
    }
    return {'total_records': total, 'status_breakdown': statuses}


@app.post('/crm/import-csv')
def import_csv(file: bytes = None, db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    from io import StringIO
    import csv

    if file is None:
        raise HTTPException(status_code=400, detail='CSV file is required')

    text = file.decode('utf-8')
    reader = csv.DictReader(StringIO(text))

    created = 0
    for row in reader:
        try:
            township = int(row.get('township', 0))
            rng = int(row.get('range', 0))
            section = int(row.get('section', 0))
        except ValueError:
            continue

        trscode = f"T{township}R{rng}S{section}"
        crm = CrmRecord(
            company=row.get('company', ''),
            contact=row.get('contact', ''),
            status=row.get('status', 'No Contact'),
            township=township,
            range=rng,
            section=section,
            trscode=trscode,
            extra_data={k: v for k, v in row.items() if k not in ['company', 'contact', 'status', 'township', 'range', 'section']}
        )
        db.add(crm)
        created += 1

    db.commit()
    return {'created': created}


def generate_pdf_report(records):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    c.drawString(100, height - 50, "CRM Records Report")
    c.drawString(100, height - 70, f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}")

    y = height - 100
    for record in records:
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(50, y, f"ID: {record.id}, Company: {record.company}, Contact: {record.contact}, Status: {record.status}")
        y -= 20

    c.save()
    buffer.seek(0)
    return buffer


@app.get('/crm/pdf-report')
def download_pdf_report(db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    records = db.query(CrmRecord).limit(100).all()
    pdf_buffer = generate_pdf_report(records)
    return FileResponse(pdf_buffer, media_type='application/pdf', filename='crm_report.pdf')


@app.on_event('startup')
def seed_admin():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.is_admin == True).first()
        if not admin:
            admin_user = User(
                email='admin@localerp.com',
                hashed_password=get_password_hash('admin'),
                is_active=True,
                is_admin=True,
                is_manager=True,
            )
            db.add(admin_user)
            db.commit()
            print(">>> Default admin created: admin@localerp.com / admin")
    finally:
        db.close()


@app.get('/')
def serve_frontend():
    return FileResponse(FRONTEND_DIR / "index.html")


app.mount('/frontend', StaticFiles(directory=str(FRONTEND_DIR)), name='frontend')


